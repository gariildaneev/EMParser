import json
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from src.logger.logger import parser_logger


class ExcelSaver:
    # Add a class-level flag to track if the file has been cleaned
    _is_file_cleaned = False

    def __init__(self, excel_file="output.xlsx", json_folder="json_data", articles=None):
        """Инициализирует объект для работы с Excel и JSON-данными."""
        
        try:
            parser_logger.info(f"{self.__class__.__name__}: Инициализация класса")

            self.excel_file = excel_file  # Фиксированный путь к Excel-файлу
            self.json_folder = json_folder
            self.articles = articles or []  # Список артикулов передаётся из GUI
            self.workbook = None  # Workbook для работы с несколькими листами

            parser_logger.debug(
                f"{self.__class__.__name__}: Экземпляр создан с excel_file='{self.excel_file}', json_folder='{self.json_folder}', articles={len(self.articles)}")

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при инициализации класса: {e}")

    def _get_latest_json(self, folder):
        """Находит самый свежий JSON-файл в указанной папке."""
        
        try:
            parser_logger.info(f"{self.__class__.__name__}: Поиск самого свежего JSON-файла в папке '{folder}'")

            # Получаем список JSON-файлов
            json_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".json")]
            parser_logger.debug(f"{self.__class__.__name__}: Найдено {len(json_files)} JSON-файлов: {json_files}")

            # Если файлов нет — выдаем предупреждение и исключение
            if not json_files:
                parser_logger.warning(
                    f"{self.__class__.__name__}: Нет JSON-файлов в папке '{folder}', поиск невозможен")
                raise FileNotFoundError("Нет JSON-файлов в указанной папке.")

            # Определяем самый свежий файл
            latest_file = max(json_files, key=os.path.getmtime)
            parser_logger.info(f"{self.__class__.__name__}: Выбран самый свежий JSON-файл: '{latest_file}'")

            return latest_file

        except FileNotFoundError:
            parser_logger.warning(f"{self.__class__.__name__}: Ошибка: Папка '{folder}' не найдена или пустая")
            raise

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при поиске JSON-файлов в '{folder}': {e}")
            raise

    def _load_price_from_json(self):
        """Загружает данные из JSON-файла."""
        
        try:
            self.json_file = self._get_latest_json(self.json_folder)  # Выбираем самый свежий JSON-файл
            parser_logger.info(f"{self.__class__.__name__}: Начало загрузки данных из {self.json_file}")

            # Читаем JSON
            with open(self.json_file, 'r', encoding='utf-8') as file:
                data = json.load(file)

            parser_logger.info(
                f"{self.__class__.__name__}: Данные из {self.json_file} успешно загружены ({len(data)} записей)")
            return data

        except FileNotFoundError:
            parser_logger.warning(
                f"{self.__class__.__name__}: Файл JSON не найден ({self.json_file}), данные не загружены")
            return {}

        except json.JSONDecodeError as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при разборе JSON-файла {self.json_file}: {e}")
            return {}

    def _open_excel(self):
        """Открывает существующий Excel-файл."""
        
        try:
            parser_logger.info(f"{self.__class__.__name__}: Открытие Excel-файла '{self.excel_file}'")

            # Загружаем книгу
            self.workbook = load_workbook(self.excel_file)

        except FileNotFoundError:
            parser_logger.warning(f"{self.__class__.__name__}: Файл '{self.excel_file}' не найден, процесс остановлен")
            raise  # Останавливаем выполнение

        except Exception as e:
            parser_logger.exception(
                f"{self.__class__.__name__}: Ошибка при загрузке Excel-файла '{self.excel_file}': {e}")
            raise

    def _create_json_sheet(self):
        """Создаёт новый лист с именем JSON-файла и записывает данные по артикулам в заданном формате."""
        
        try:
            self.json_file = self._get_latest_json(self.json_folder)
            parser_logger.info(f"{self.__class__.__name__}: Создание нового листа из JSON-файла '{self.json_file}'")

            self.data = self._load_price_from_json()
            sheet_name = os.path.basename(self.json_file).split('.')[0].split("_")[0].removesuffix('Data')
            parser_logger.debug(f"{self.__class__.__name__}: Имя нового листа: {sheet_name}")

            if sheet_name in self.workbook.sheetnames:
                parser_logger.warning(f"{self.__class__.__name__}: Лист '{sheet_name}' уже существует, удаляем его")
                self.workbook.remove(self.workbook[sheet_name])

            self.workbook.create_sheet(sheet_name)
            ws = self.workbook[sheet_name]

            current_col = 1  # Начальная колонка
            for article in self.articles:
                col_letter = get_column_letter(current_col)
                next_col_letter = get_column_letter(current_col + 1)
                link_col_letter = get_column_letter(current_col + 2)

                # Объединяем три колонки для артикула
                ws.merge_cells(f"{col_letter}1:{link_col_letter}1")
                ws[f"{col_letter}1"] = article
                ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

                # Записываем заголовки "name", "price" и "link"
                ws[f"{col_letter}2"] = "name"
                ws[f"{next_col_letter}2"] = "price"
                ws[f"{link_col_letter}2"] = "link"

                row = 3  # Начальная строка для данных
                for item in self.data.get("Данные", []):
                    if article in item:
                        description = item[article].get("description", "Не найдено")
                        price = item[article].get("price", "Не найдено")
                        link = item[article].get("url", "")
                        ws[f"{col_letter}{row}"] = description
                        ws[f"{next_col_letter}{row}"] = price
                        ws[f"{link_col_letter}{row}"] = link
                        if link:  # Если есть ссылка, делаем её кликабельной
                            ws[f"{link_col_letter}{row}"].hyperlink = link
                            ws[f"{link_col_letter}{row}"].style = "Hyperlink"
                        row += 1

                current_col += 3  # Переход к следующей тройке колонок

            parser_logger.info(f"{self.__class__.__name__}: Создан новый лист '{sheet_name}' в заданном формате")

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при создании листа '{sheet_name}': {e}")

    def _save_to_excel(self):
        """Сохраняет изменения в Excel-файл."""
        
        try:
            parser_logger.info(f"{self.__class__.__name__}: Сохранение изменений в '{self.excel_file}'")

            self.workbook.save(self.excel_file)

            parser_logger.info(f"{self.__class__.__name__}: Изменения успешно сохранены в '{self.excel_file}'")

        except PermissionError:
            parser_logger.exception(
                f"{self.__class__.__name__}: Ошибка: Файл '{self.excel_file}' открыт в другой программе, сохранение невозможно")
        except Exception as e:
            parser_logger.exception(
                f"{self.__class__.__name__}: Ошибка при сохранении в Excel-файл '{self.excel_file}': {e}")

    def process_data(self):
        """Performs the full data processing cycle: loading, updating, and saving."""
        try:
            parser_logger.info(f"{self.__class__.__name__}: Начало обработки данных")

            # Clean the Excel file only if it hasn't been cleaned yet
            if not ExcelSaver._is_file_cleaned:
                parser_logger.info(f"{self.__class__.__name__}: Очистка файла '{self.excel_file}'")
                self.workbook = Workbook()  # Create a new workbook
                self.workbook.save(self.excel_file)  # Save the empty workbook
                parser_logger.info(f"{self.__class__.__name__}: Файл '{self.excel_file}' успешно очищен")
                ExcelSaver._is_file_cleaned = True  # Mark the file as cleaned

            # Open the existing Excel file
            self._open_excel()
            parser_logger.info(f"{self.__class__.__name__}: Excel-файл успешно загружен")

            # Create a new sheet with JSON data
            self._create_json_sheet()
            parser_logger.info(f"{self.__class__.__name__}: Новый лист с JSON-данными успешно создан")

            # Save changes to the Excel file
            self._save_to_excel()
            parser_logger.info(f"{self.__class__.__name__}: Изменения сохранены в Excel-файл")

            parser_logger.info(f"{self.__class__.__name__}: Обработка данных завершена")

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка во время обработки данных: {e}")

    def aggregate_prices_to_first_sheet(self):
        """Собирает все цены с каждого листа и аккумулирует их на первом листе."""
        
        try:
            parser_logger.info(f"{self.__class__.__name__}: Начало агрегации цен в Excel-файл '{self.excel_file}'")

            # Открываем Excel
            self._open_excel()
            first_sheet_name = self.workbook.sheetnames[0]
            first_sheet = self.workbook[first_sheet_name]
            parser_logger.debug(f"{self.__class__.__name__}: Первый лист для агрегации: '{first_sheet_name}'")

            # Очистка первой страницы от старых данных
            first_sheet.delete_rows(1, first_sheet.max_row)

            # Словарь для хранения данных по артикулам
            aggregated_data = {}

            for sheet_name in self.workbook.sheetnames[1:]:
                sheet = self.workbook[sheet_name]
                parser_logger.info(f"{self.__class__.__name__}: Обрабатываем лист '{sheet_name}'")

                for col in range(1, sheet.max_column + 1, 3):  # Обрабатываем тройки колонок
                    article_cell = sheet.cell(row=1, column=col)
                    if not article_cell.value:
                        continue

                    article = article_cell.value
                    shop_name = sheet_name  # Название магазина — имя листа

                    # Инициализируем список для артикула, если его ещё нет
                    if article not in aggregated_data:
                        aggregated_data[article] = []

                    # Собираем данные
                    for data_row in range(3, sheet.max_row + 1):
                        name = sheet.cell(row=data_row, column=col).value
                        price = sheet.cell(row=data_row, column=col + 1).value
                        link = sheet.cell(row=data_row, column=col + 2).value
                        if name or price:  # Добавляем только если есть данные
                            aggregated_data[article].append({
                                "shop": shop_name,
                                "name": name,
                                "price": price,
                                "link": link
                            })

            # Записываем данные на первый лист
            current_col = 1
            for article, records in aggregated_data.items():
                # Объединяем четыре колонки для артикула
                first_sheet.merge_cells(
                    start_row=1, start_column=current_col, end_row=1, end_column=current_col + 3
                )
                first_sheet.cell(row=1, column=current_col).value = article
                first_sheet.cell(row=1, column=current_col).alignment = Alignment(horizontal="center", vertical="center")

                # Записываем заголовки "shop", "name", "price" и "link"
                first_sheet.cell(row=2, column=current_col).value = "shop"
                first_sheet.cell(row=2, column=current_col + 1).value = "name"
                first_sheet.cell(row=2, column=current_col + 2).value = "price"
                first_sheet.cell(row=2, column=current_col + 3).value = "link"

                # Записываем данные
                row = 3
                for record in records:
                    first_sheet.cell(row=row, column=current_col).value = record["shop"]
                    first_sheet.cell(row=row, column=current_col + 1).value = record["name"]
                    first_sheet.cell(row=row, column=current_col + 2).value = record["price"]
                    link_cell = first_sheet.cell(row=row, column=current_col + 3)
                    link_cell.value = record["link"]
                    if link_cell.value:  # Если есть ссылка, делаем её кликабельной
                        link_cell.hyperlink = link_cell.value
                        link_cell.style = "Hyperlink"
                    row += 1

                current_col += 4  # Переход к следующей четвёрке колонок

            parser_logger.info(f"{self.__class__.__name__}: Агрегация завершена, данные сохранены в '{self.excel_file}'")

            # Сохранение изменений
            self._save_to_excel()

        except Exception as e:
            parser_logger.exception(f"{self.__class__.__name__}: Ошибка при агрегации цен: {e}")


